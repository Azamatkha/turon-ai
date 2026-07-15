"""Parse a bank-directory Excel (.xlsx) into employee records.

Each worksheet is one department (sheet title = department name). A sheet has a
title row (or a few) at the top, a header row — detected by CONTENT, not a fixed
index, because the header may span several rows and start at a different row in
different sheets — then data rows. Some data rows are section dividers (a
sub-division name with no phone/IP); they set the "division" for the employees
that follow.
"""

from io import BytesIO
from typing import Any


def _txt(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: Any) -> str:
    return _txt(value).lower().replace(" ", "").replace(".", "")


def _find_header(
    rows: list[tuple[Any, ...]],
) -> tuple[int, dict[str, int]] | None:
    """Header qatorini (matn bo'yicha) topib, mantiqiy ustunlarni indeksga
    moslaydi. Topilmasa None."""
    def is_fish(s: str) -> bool:
        return "fish" in s or "fio" in s or "фиш" in s or "фио" in s

    def is_phone(s: str) -> bool:
        return "telefon" in s or "телефон" in s

    def is_ip(s: str) -> bool:
        return "ichkiraqam" in s or "(ip)" in s or "ip)" in s or "ички" in s

    def is_pos(s: str) -> bool:
        return "tuzilma" in s or "lavozim" in s or "тузилма" in s or "лавозим" in s

    for i, row in enumerate(rows):
        joined = "".join(_norm(c) for c in row)
        if not (is_fish(joined) and is_ip(joined) and is_phone(joined)):
            continue
        col: dict[str, int] = {}
        for j, c in enumerate(row):
            cn = _norm(c)
            if not cn:
                continue
            if is_fish(cn) and "fish" not in col:
                col["fish"] = j
            elif is_phone(cn) and "phone" not in col:
                col["phone"] = j
            elif is_ip(cn) and "ip" not in col:
                col["ip"] = j
            elif is_pos(cn) and "position" not in col:
                col["position"] = j
        return i, col
    return None


def parse_employees(file_bytes: bytes) -> list[dict[str, str]]:
    """Excel baytlaridan xodimlar ro'yxatini qaytaradi.
    Har biri: {department, division, position, fish, ip, phone}."""
    # Lazy import — openpyxl o'rnatilmagan bo'lsa ilova ishga tushishi buzilmasin
    # (faqat shu funksiya chaqirilganda xato beradi).
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    records: list[dict[str, str]] = []
    try:
        for ws in wb.worksheets:
            department = _txt(ws.title)
            rows = list(ws.iter_rows(values_only=True))
            found = _find_header(rows)
            if not found:
                continue
            header_idx, col = found
            if "fish" not in col or "ip" not in col:
                continue

            division = ""
            for row in rows[header_idx + 1 :]:
                if not any(_txt(c) for c in row):
                    continue

                def cell(key: str) -> str:
                    idx = col.get(key, -1)
                    return _txt(row[idx]) if 0 <= idx < len(row) else ""

                fish = cell("fish")
                ip = cell("ip")
                phone = cell("phone")
                position = cell("position")

                # Xodim qatorida ism yoki aloqa raqami bo'ladi. Aks holda bu —
                # ichki bo'linma (division) sarlavhasi: keyingi xodimlar uchun
                # eslab qolamiz.
                if not fish and not ip and not phone:
                    texts = [_txt(c) for c in row if _txt(c)]
                    if texts:
                        division = max(texts, key=len)
                    continue

                records.append(
                    {
                        "department": department,
                        "division": division,
                        "position": position,
                        "fish": fish,
                        "ip": ip,
                        "phone": phone,
                    }
                )
    finally:
        wb.close()
    return records
